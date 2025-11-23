#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для проверки статуса авиакомпаний
Проверяет: актуальность, статус операций, переименование
"""

import requests
import time
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import json

class AirlineChecker:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
    def check_wikipedia(self, airline_name):
        """Проверка через Wikipedia API"""
        try:
            # Поиск статьи
            search_url = "https://en.wikipedia.org/w/api.php"
            params = {
                'action': 'opensearch',
                'search': airline_name,
                'limit': 5,
                'format': 'json'
            }
            
            response = self.session.get(search_url, params=params, timeout=10)
            results = response.json()
            
            if len(results) > 1 and len(results[1]) > 0:
                # Получаем содержимое первой найденной статьи
                article_title = results[1][0]
                content_params = {
                    'action': 'query',
                    'titles': article_title,
                    'prop': 'extracts',
                    'exintro': True,
                    'explaintext': True,
                    'format': 'json'
                }
                
                content_response = self.session.get(search_url, params=content_params, timeout=10)
                content_data = content_response.json()
                
                pages = content_data.get('query', {}).get('pages', {})
                if pages:
                    page = list(pages.values())[0]
                    extract = page.get('extract', '').lower()
                    
                    return {
                        'found': True,
                        'title': article_title,
                        'url': results[2][0] if len(results) > 2 else '',
                        'extract': extract
                    }
            
            return {'found': False}
            
        except Exception as e:
            print(f"Wikipedia error for {airline_name}: {str(e)}")
            return {'found': False, 'error': str(e)}
    
    def analyze_status(self, text, airline_name):
        """Анализ статуса авиакомпании по тексту"""
        text_lower = text.lower()
        airline_lower = airline_name.lower()
        
        # Ключевые слова для определения статуса
        defunct_keywords = [
            'ceased operations', 'defunct', 'no longer operates', 
            'discontinued', 'liquidated', 'bankrupt', 'shut down',
            'stopped flying', 'ended operations', 'closed down',
            'ceased trading', 'went out of business'
        ]
        
        operating_keywords = [
            'currently operates', 'operating', 'operates flights',
            'active airline', 'continues to operate', 'flying',
            'serves destinations', 'scheduled flights', 'is operating'
        ]
        
        renamed_keywords = [
            'renamed to', 'rebranded as', 'now known as',
            'changed its name to', 'became', 'merged with',
            'acquired by', 'replaced by'
        ]
        
        # Проверка статуса
        is_defunct = any(keyword in text_lower for keyword in defunct_keywords)
        is_operating = any(keyword in text_lower for keyword in operating_keywords)
        is_renamed = any(keyword in text_lower for keyword in renamed_keywords)
        
        # Поиск даты прекращения операций
        ceased_date_pattern = r'ceased operations?.*?(\d{4})'
        ceased_match = re.search(ceased_date_pattern, text_lower)
        ceased_year = ceased_match.group(1) if ceased_match else None
        
        # Поиск нового названия
        new_name = None
        if is_renamed:
            for keyword in renamed_keywords:
                pattern = f'{keyword}\\s+([A-Z][\\w\\s&-]+?)(?:\\.|,|\\sin\\s|\\sfrom\\s|$)'
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    new_name = match.group(1).strip()
                    break
        
        # Определение уровня уверенности
        confidence = 'НИЗКИЙ'
        if is_defunct and ceased_year:
            confidence = 'ВЫСОКИЙ'
        elif is_operating and 'currently' in text_lower:
            confidence = 'ВЫСОКИЙ'
        elif is_defunct or is_operating:
            confidence = 'СРЕДНИЙ'
        
        # Определение статуса
        if is_defunct:
            status = f'НЕ ДЕЙСТВУЕТ (прекращена {ceased_year if ceased_year else "дата неизвестна"})'
        elif is_operating:
            status = 'ДЕЙСТВУЕТ'
        elif is_renamed and new_name:
            status = f'ПЕРЕИМЕНОВАНА'
        else:
            status = 'СТАТУС НЕИЗВЕСТЕН'
        
        return {
            'status': status,
            'new_name': new_name if new_name else 'Н/Д',
            'confidence': confidence,
            'ceased_year': ceased_year
        }
    
    def check_airline(self, airline_name):
        """Комплексная проверка авиакомпании"""
        print(f"Проверка: {airline_name}")
        
        # Проверка Wikipedia
        wiki_result = self.check_wikipedia(airline_name)
        
        if not wiki_result.get('found'):
            return {
                'airline': airline_name,
                'status': 'ИНФОРМАЦИЯ НЕ НАЙДЕНА',
                'new_name': 'Н/Д',
                'confidence': 'НИЗКИЙ',
                'source': 'Информация не найдена в доступных источниках'
            }
        
        # Анализ найденной информации
        analysis = self.analyze_status(wiki_result['extract'], airline_name)
        
        return {
            'airline': airline_name,
            'status': analysis['status'],
            'new_name': analysis['new_name'],
            'confidence': analysis['confidence'],
            'source': f"Wikipedia: {wiki_result['url']}"
        }
    
    def create_excel_report(self, results, filename='airline_status_report.xlsx'):
        """Создание Excel отчета"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Статус авиакомпаний"
        
        # Заголовки
        headers = ['№', 'Название авиакомпании', 'Статус', 'Новое название (если переименована)', 
                   'Уровень уверенности', 'Источник информации']
        
        # Стили для заголовков
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 50
        
        # Цвета для уровней уверенности
        confidence_colors = {
            'ВЫСОКИЙ': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'СРЕДНИЙ': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'НИЗКИЙ': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        }
        
        # Заполнение данных
        for idx, result in enumerate(results, 2):
            ws.cell(row=idx, column=1, value=idx-1)
            ws.cell(row=idx, column=2, value=result['airline'])
            ws.cell(row=idx, column=3, value=result['status'])
            ws.cell(row=idx, column=4, value=result['new_name'])
            
            # Ячейка с уровнем уверенности
            confidence_cell = ws.cell(row=idx, column=5, value=result['confidence'])
            if result['confidence'] in confidence_colors:
                confidence_cell.fill = confidence_colors[result['confidence']]
            confidence_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.cell(row=idx, column=6, value=result['source'])
            
            # Выравнивание
            for col in range(1, 7):
                ws.cell(row=idx, column=col).alignment = Alignment(vertical='center', wrap_text=True)
        
        # Закрепление первой строки
        ws.freeze_panes = 'A2'
        
        # Добавление информационного листа
        info_ws = wb.create_sheet('Информация')
        info_data = [
            ['Отчет о проверке статуса авиакомпаний'],
            ['Дата создания:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Всего проверено:', len(results)],
            [''],
            ['Легенда уровней уверенности:'],
            ['ВЫСОКИЙ', 'Информация подтверждена из надежных источников, найдены конкретные даты и факты'],
            ['СРЕДНИЙ', 'Информация найдена, но требует дополнительной проверки'],
            ['НИЗКИЙ', 'Информация не найдена или противоречива'],
            [''],
            ['Статусы:'],
            ['ДЕЙСТВУЕТ', 'Авиакомпания выполняет регулярные рейсы'],
            ['НЕ ДЕЙСТВУЕТ', 'Авиакомпания прекратила операции'],
            ['ПЕРЕИМЕНОВАНА', 'Авиакомпания изменила название'],
            ['ИНФОРМАЦИЯ НЕ НАЙДЕНА', 'Не удалось найти информацию в доступных источниках']
        ]
        
        for row_idx, row_data in enumerate(info_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                info_ws.cell(row=row_idx, column=col_idx, value=value)
        
        info_ws.column_dimensions['A'].width = 30
        info_ws.column_dimensions['B'].width = 80
        
        wb.save(filename)
        print(f"\nОтчет сохранен: {filename}")
        return filename

def main():
    # Список авиакомпаний
    airlines = """Aegean Airlines
Aer Lingus
Aereonautica militare
Aero Contractors
Aero Lanka
Aero VIP (2D)
Aero-Service
Aeroflot Russian Airlines
Aerolane
Aeroline GmbH
Aerolineas Africanas
Aerolineas Argentinas
Aerolineas Galapagos (Aerogal)
Aerolineas heredas santa maria
Aerolitoral
Aeromar
AeroMéxico
Aeronautica Militare
Aeronorte
Aeropelican Air Services
AeroRep
Aerosur
AeroWorld 
Africa West
Afriqiyah Airways
Air 26
Air Afrique
Air Alfa
Air Algerie
Air Alps Aviation (A6)
Air Antilles Express
Air Arabia
Air Arabia Egypt
Air Arabia Maroc
Air Astana
Air Atlanta Icelandic
Air Austral
Air Bagan
Air Baltic
Air Bangladesh
Air Batumi
Air Botswana
Air Bucharest
Air Burkina
Air Busan
Air Caledonie International
Air Canada
Air Canada Jazz
Air Caraïbes
Air Cargo Carriers
Air Cargo Germany
Air Carnival
Air Century
Air Chathams
Air China
Air Choice One
Air Comet Chile
Air Costa
Air Cudlua
Air Dolomiti
Air Europa
Air Exel
Air Explore
Air Finland
Air Florida
Air Foyle
Air France
Air Glaciers
Air Greenland
Air Guinee Express
Air Guyane
Air Hamburg (AHO)
Air Iceland
Air India Express
Air India Limited
Air India Regional
AIR INDOCHINE
Air Indus
Air Italy Egypt
Air Ivoire
Air Japan
Air Kazakhstan
Air KBZ
Air Kenya (Priv)
Air Kiribati
Air Koryo
Air Lituanica
Air Luxor
Air Macau
Air Madagascar
Air Madrid
Air Majoro
Air Malta
Air Mandalay
Air Marshall Islands
Air Mauritius
Air Mediterranee
Air Mekong
Air Moldova
Air Moorea
Air Mozambique
Air New Zealand
Air Niugini
Air North Charter - Canada
Air Nostrum
Air Pacific
Air Pegasus
Air Philippines
Air Plus Comet
Air Rarotonga
Air Saint Pierre
Air Serbia
Air Seychelles
Air Sinai
Air Sunshine
Air Tahiti
Air Tahiti Nui
Air Tanzania
Air Tindi
Air Transat
Air Vanuatu
Air VIA
Air Vistara
Air Volga
Air Wales
Air Wisconsin
Air Zimbabwe
Air2there
AirAsia
AirAsia Japan
AirAsia X
Airblue
Aircompany Yakutia
Aires
Airfix Aviation
Airlift International
Airlinair
Airlines PNG
Airlink (SAA)
Airnorth
AirOnix
AirRussia
AirTran Airways
Akasa Air
Al-Naser Airlines
ALAK
Alaska Airlines
Alaska Seaplane Service
Albanian Airlines
AlbaStar
Alghanim
Alitalia
All Africa
All America
All America AR
All America BOPY
All America BR
All America CL
All America CO
All America MX
All America US
All Argentina
All Argentina Express
All Asia
All Australia
All Colombia
All Europe
All Nippon Airways
All Spain
Allegiant Air
Alliance Airlines
AlMasria Universal Airlines
Alrosa Mirny Air Enterprise
Amaszonas
American Airlines
American Eagle Airlines
Amsterdam Airlines
ANA Wings
Andalus Lineas Aereas
Apache Air
Ariana Afghan Airlines
Arik Niger
Arkefly
Arkia Israel Airlines
Armavia
Arrow Air
Aruba Airlines
Aserca Airlines
Asia Wings
Asian Spirit
Asian Wings Airways
Asiana Airlines
Askari Aviation
Astair
Astral Aviation
Athens Airways
Atifly
Atlant-Soyuz Airlines
Atlantic Airways
Atlantis European Airways
Atlas Air
AtlasGlobal Ukraine
Atlasjet
ATRAN Cargo Airlines
Aurigny Air Services
Austral Brasil
Austral Lineas Aereas
Austrian Airlines
Aviabus
Avialeasing Aviation Company
Avianca
Avianca - Aerovias Nacionales de Colombia
Avianova (Russia)
Avient Aviation
Avilu
Aws express
Azerbaijan Airlines
Azul
BA CityFlyer
BA101
Bahamasair
Baikotovitchestrian Airlines 
Balkan Bulgarian Airlines
Baltic Air lines
Baltic Airlines
Bamboo Airways
Bangkok Airways
Barents AirLink
Bassaka airlines
Bateleur Air
Batik Air
BBN-Airways
Bearskin Lake Air Service
Belair Airlines
Belavia Belarusian Airlines
Bellview Airlines
Bemidji Airlines
Bering Air
Berjaya Air
BF-Lento OY
Biman Bangladesh Airlines
Bingo Airways
Binter Canarias
Black Stallion Airways
Blue Air
Blue Panorama Airlines
Blue Wings
Blue1
Bluebird Airways (BZ)
bmi
bmibaby
Bobb Air Freight
Boliviana de Aviacion (OB)
Boutique Air (Priv)
BQB Lineas Aereas
BRAZIL AIR
Brazilian Air Force
Breeze Airways
Brit Air
British Airways
British International Helicopters
British Mediterranean Airways
British Midland Regional
Brussels Airlines
Budapest Aircraft Services/Manx2
BudgetAir
Bulgaria Air
Buquebus Líneas Aéreas
Business Aviation
BusinessAir
BVI Airways
CAL Cargo Air Lines
Calima Aviacion
Camair-co
Cambodia Angkor Air (K6)
Canadian Airlines
Canadian National Airways
Canadian North
Canaryfly
CanXplorer
CanXpress
Cape Air
Cargo Plus Aviation
Caribbean Airlines
CARICOM AIRWAYS (BARBADOS) INC.
Carnival Air Lines
Carpatair
Carpatair Flight Training
Caspian Airlines
Cathay Pacific
Catovair
Caucasus Airlines
Cayman Airways
CB Airways UK ( Interliging Flights )
CBM America
CCML Airlines
Cebu Pacific
CEIBA Intercontinental
Cello Aviation
Central Connect Airlines
Centralwings
Charter Air
Chautauqua Airlines
China Airlines
China Eastern Airlines
China Northwest Airlines (WH)
China Southern Airlines
China SSS
China United Airlines
Choice Airways
Chongqing Airlines
Ciao Air
Ciel Canadien
Cimber Air
Cirrus Airlines
Citilink Indonesia
City Airline
City Airways
City Connexion Airlines
CityJet
Click (Mexicana)
Click Airways
Coastal Air
COBALT
Colgan Air
Comair
Comfort Express Virtual Charters
Comfort Express Virtual Charters Albany
CommutAir
Comores Airlines
Compagnie Africaine d\\'Aviation
Compass Airlines
Comtel Air
Condor Flugdienst
Congo Express
ConneX European Airline
Consorcio Aviaxsa
Contact Air
Continental Airlines
Continental Express
Continental Micronesia
Conviasa
Copa Airlines
Copenhagen Express
Copterline
Corendon Airlines
Corsairfly
Corse-Mediterranee
Crest Aviation
Croatia Airlines
Crown Airways
Cubana de Aviación
Cyprus Airways
Cyprus Turkish Airlines
Czech Airlines
Daallo Airlines
Dalavia
Dana Air
Danube Wings (V5)
Darwin Airline
DAT Danish Air Transport
dba
Delta Aerotaxi
Delta Air Lines
Denim Air
Dennis Sky
Dense Airways
Dense Connection
Deutsche Bahn
Direct Aero Services
Djibouti Airlines
Dniproavia
Dobrolet
Domenican Airlines
Dominicana de Aviaci
Domodedovo Airlines
DonbassAero
Dragonair
Druk Air
Dubrovnik Air
Dutch Antilles Express
Dynamic Airways
Eagle Air
Eagle Atlantic Airlines
Eagles Airlines
East African
East Horizon
Eastar Jet
Eastern Airways
Eastern Atlantic Virtual Airlines
Eastland Air
Eastok Avia
EasyFly
easyJet
EasyJet (DS)
Ecuavia
Edelweiss Air
Egyptair
EgyptAir Express
El Al Israel Airlines
El-Buraq Air Transport
ELK Airways
Elysian Airlines
Emirates
Empresa Ecuatoriana De Aviacion
Enerjet
ENTERair
Envoy Air
Epic Holiday
Era Alaska
Eritrean Airlines
Estonian Air
Ethiopian Airlines
Etihad Airways
Euro Exec Express
Euro Jet
Eurocypria Airlines
Eurofly Service
Euroline
Eurolot
Europe Jet
European Air Express
EuropeSky
Eurowings
EVA Air
Evergreen International Airlines
Excel Airways
Excel Charter
Executive AirShare
Express One International
ExpressJet
Eznis Airways
Far Eastern Air Transport
Fars Air Qeshm
Feeder Airlines
Felix Airways
Finlandian
Finnair
Finncomm Airlines
Firefly
First Air
First Choice Airways
First Flying
Flightline
Flightlink Tanzania
Florida West International Airways
Flugfelag Vestmannaeyja
Fly 6ix
Fly Africa Zimbabwe
Fly Brasil
Fly Colombia ( Interliging Flights )
Fly Dubai
Fly Europa
Fly France
Fly Jamaica Airways
Fly One
Fly Romania
Fly540
Flybaboo
Flybe
Flybe Finland Oy
FlyEgypt
Flyglobespan
FlyHigh Airlines Ireland (FH)
Flyhy Cargo Airlines
Flying kangaroo Airline
FLYJET
FlyLal
FlyLAL Charters
Flyme (VP)
FlyNordic
FlyPortugal
Formosa Airlines
FOX Linhas Aereas
Freedom Air
Frontier Airlines
Frontier Flying Service
Fuji Dream Airlines
Fuzhou Airlines
Gabon Airlines
Gadair European Airlines
Garuda Indonesia
Gazpromavia
GB Airways
Genesis
Georgian Airways
Georgian International Airlines
Georgian National Airlines
German Air Force - FLB
German International Air Lines
Germania
Germanwings
GermanXL
Ghana International Airlines
Global Airlines
Globus
GNB Linhas Aereas
Go Air
Go First
Go2Sky
GoJet Airlines
Gol Transportes Aéreos
Golden Air
Golden Myanmar Airlines
Gomelavia
Grand China Air
Grand Cru Airlines
Grant Aviation
Great Lakes Airlines
Greenfly
Grupo TACA
Gryphon Airlines
Gulf Air
Gulf Air Bahrain
Gulfstream International Airlines
GX Airlines
Hageland Aviation Services
Hainan Airlines
Haiti Ambassador Airlines
Halcyonair
Hamburg Airways
Hamburg International
Hankook Air US
Hankook Airline
Hapagfly
Happy Air
Harbour Air (Priv)
Hawaiian Airlines
Hawkair
Hebei Airlines
Hebradran Air Services
Heli France
Helijet
Helitt Líneas Aéreas
Hellas Jet
Hellenic Imperial Airways
Hello
Helvetic Airways
Hermes Airlines
Hex'Air
Hi Fly (5K)
Highland Airways
Himalayan Airlines
Hokkaido International Airlines
Holidays Czech Airlines
Homer Air
Hong Kong Airlines
Hong Kong Express Airways
Horizon Air
Huaxia
I-Fly
Iberia Airlines
Iberia Express
Iberworld
Ibex Airlines
Icar Air
Icelandair
Imair Airlines
INAVIA Internacional
Indian Airlines
IndiGo Airlines
Indochina Airlines
Indonesa Air Aisa X
Indonesia AirAsia
Indonesia Sky
Indonesian Airlines
Indya Airline Group
Insel Air (7I/INC) (Priv)
Interair South Africa
Interavia Airlines
Interjet (ABC Aerolineas)
Interlink Airlines
International AirLink
International Europe
Intersky
Iran Air
Iran Aseman Airlines
Iraqi Airways
Island Air (WP)
Island Airlines
Island Express Air
Island Spirit
Islas Airways
Islena De Inversiones
Israir
ITA Airways
Itek Air
IzAvia
Jagson Airlines
JAL Express
JALways
Japan Air System
Japan Airlines
Japan Airlines Domestic
Japan Asia Airways
Japan Regio
Japan Transocean Air
Jayrow
Jazeera Airways
Jc royal.britannica
Jeju Air
Jet Airways
Jet Suite
Jet2.com
Jet4You
Jetairfly
JetBlue Airways
Jetflite
Jetgo Australia
Jetstar Airways
Jetstar Asia Airways
Jetstar Japan 
Jetstar Pacific
Jettor Airlines
Jin Air
JobAir
Jota Aviation
Joy Air
Juneyao Airlines
Jupiter Airlines
Jusur airways
Kal Star Aviation
Kam Air
Kan Air
Kavminvodyavia
KD Avia
Kendell Airlines
Kenmore Air
Kenya Airways
Kharkiv Airlines
Kingfisher Airlines
Kinloss Flying Training Unit
Kish Air
KLM Cityhopper
KLM Royal Dutch Airlines
KMV
Kogalymavia Air Company
KoralBlue Airlines
Korea Express Air
Korean Air
Korongo Airlines
Kosmos
Kostromskie avialinii
Krasnojarsky Airlines
KSY
Kuban Airlines
Kush Air
Kuwait Airways
Kuzu Airlines Cargo
L
LACSA
LAN Airlines
LAN Argentina
LAN Express
LAN Peru
Lao Airlines
LatCharter
Lauda Air
LCM AIRLINES
Leeward Islands Air Transport
Liberty Airways
Libyan Arab Airlines
Line Blue
Linhas A
Lion Mentari Airlines
LionXpress
Locair
LOT Polish Airlines
LSM Airlines
LSM International 
LTE International Airways
LTU Austria
Luchsh Airlines 
Luftfahrtgesellschaft Walter
Lufthansa
Lufthansa Cargo
Lufthansa CityLine
Lufttransport
Lugansk Airlines
Luxair
Maastricht Airlines
Macair Airlines
Maersk
Mahan Air
Malawian Airlines
Malaysia Airlines
Malaysia Wings
Maldivian
Maldivian Air Taxi
Maldivo Airlines
Malindo Air
Malmo Aviation
Malmö Aviation
Malta Air Charter
Malév
Mandala Airlines
Mandarin Airlines
Mango
Mann Yadanarpon Airlines
Martinair
Marusya Airways
Maryland Air
Marysya Airlines
MasAir
MAT Airways
MAT Macedonian Airlines
Mauritania Airlines International
Mauritania Airways
Maxair
Maya Island Air
MCA Airlines
MDLR Airlines
Med Airways
Medallion Air
Meridiana
Merpati Nusantara Airlines
Mesa Airlines
Mesaba Airlines
Metro Batavia
Mexicana de Aviaci
MexicanaLink
MHS Aviation GmbH
MIAT Mongolian Airlines
Michael Airlines
Middle East Airlines
Midway Airlines
Midwest Airlines
Midwest Airlines (Egypt)
Mint Airways
MNG Airlines
Moldavian Airlines
Monarch Airlines
Mongolian International Air Lines 
Montenegro Airlines
Morningstar Air Express
Moskovia Airlines
Motor Sich
Myanma Airways
Myanmar Airways International
Myflug
MyTravel Airways
Myway Airlines
N1
Nas Air
Nasair
National Air Cargo
National Jet Systems
Nationwide Airlines
Nauru Air Corporation
Nepal Airlines
Nesma Airlines
NetJets
New England Airlines
NEXT Brasil
NextJet
Nihon.jet connect
Niki
Nile Air
Nok Air
Nordic Global Airlines
Nordica
NordStar Airlines
Norfolk County Flight College
Norlandair
Norte Lineas Aereas
North American Airlines
North American Charters
Northern Dene Airways
Northwest Airlines
Northwestern Air
Norwegian Air International (D8)
Norwegian Air Shuttle
Norwegian Aviation College
Norwegian Long Haul AS
Nouvel Air Tunisie
Novair
Oasis Hong Kong Airlines
Ocean Air
OCEAN AIR CARGO
Oceanair
Oceanic Airlines
Olympic Airlines
Oman Air
Omni Air International
One Two Go Airlines
OneChina
OneJet
Onur Air
Orbest
Orbit Airlines
Orbit Airlines Azerbaijan
Orbit Atlantic Airways
Orbit International Airlines
Orbit Regional Airlines
Orchid Airlines
Orenburg Airlines
Orenburzhie
Orient Thai Airlines
Origin Pacific Airways
Ostfriesische Lufttransport
Overland Airways
Ozjet Airlines
Pacific Coastal Airline
Pacific East Asia Cargo Airlines
Pacific Flier
Pacific Island Aviation
Pacific Wings
Pakistan International Airlines
Pal airlines
PAN Air
Pan Am World Airways Dominicana
PanAm World Airways
Papillon Grand Canyon Helicopters
Paramount Airways
Parmiss Airlines (IPV)
Passaredo Transportes Aereos
Patriot Airways
PB Air
Peach Aviation
Pegasus Airlines
PEGASUS AIRLINES-
Pelita
Peninsula Airways
Peruvian Airlines
Philippine Airlines
Piedmont Airlines (1948-1989)
Pinnacle Airlines
PLUNA
PMTair
Pobeda
Polar Airlines
Polet
Polet Airlines (Priv)
Polynesian Airlines
Porter Airlines
Portugalia
Potomac Air
Pouya Air
Precision Air
Primera Air
Privatair
Proflight Commuter Services
Qantas
Qatar Airways
QatXpress
RACSA
Rainbow Air (RAI)
Rainbow Air Canada
Rainbow Air Euro
Rainbow Air Polynesia
Rainbow Air US
Real Tonga
Red Jet Andes
Red Jet Canada
Red Jet Mexico
Red Wings
REDjet
Regional Air Iceland
Regional Airlines
Regional Express
Regional Paraguaya
Regionalia Chile
Regionalia México
Regionalia Uruguay
Regionalia Venezuela
Republic Airlines
Republic Express Airlines
REXAIR VIRTUEL
Rossiya
Rossiya-Russian Airlines
Rotana Jet
Royal Air Cambodge
Royal Air Maroc
Royal Airways
Royal Brunei Airlines
Royal European Airlines
Royal Falcon
Royal Flight
Royal Jordanian
Royal Nepal Airlines
Royal Phnom Penh Airways
RusJet
Rusline
Russia State Transport
Russkie Krylya
Rwandair Express
Ryan Air Services
Ryan International Airlines
Ryanair
Régional
S7 Airlines
Safi Airlines
Sahara Airlines
Salmon Air
Salsa d\\'Haiti
Salzburg arrows
Sama Airlines
San Dima Air
Santa Barbara Airlines
Saratov Aviation Division
SAS Braathens
Sat Airlines
SATA Air Acores
SATA International
SATENA
Saudi Arabian Airlines
Scandinavian Airlines System
Scat Air
Scenic Airlines
Scoot
ScotAirways
Seaborne Airlines
SeaPort Airlines
Senegal Airlines
SENIC AIRLINES
Serbian Airlines
Servicios de Transportes A
Sevenair
Severstal Air Company
SGA Airlines
Shaheen Air International
Shandong Airlines
Shanghai Airlines
Sharp Airlines
Shenzhen Airlines
Shuttle America
Sibaviatrans
Sichuan Airlines
Siem Reap Airways
SilkAir
Silver Airways (3M)
Simrik Airlines
Singapore Airlines
Singapore Airlines Cargo
Skagway Air Service
Sky Airline
Sky Angkor Airlines (ZA)
Sky Europe Airlines
Sky Express
Sky Regional
Sky Wing Pacific
SkyAlps
SkyBahamas Airlines
SkyEurope
Skyjet Airlines
Skyline Ulasim Ticaret A.S.
Skymark Airlines
Skynet Airlines
Skynet Asia Airways
Skyservice Airlines
Skywalk Airlines
Skyways Express
Skywest Airlines
Skywest Australia
SkyWork Airlines 
Small Planet Airlines
SmartLynx Airlines
Snowbird Airlines
SOCHI AIR
SOCHI AIR CHATER
SOCHI AIR EXPRESS
Sol Lineas Aereas
Solar Air
Solomon Airlines
South African Airways
South East Asian Airlines
South Pacific Island Airways
Southeast Air
Southern Air Charter
Southern Airways
Southern Airways Express
Southern Winds Airlines
Southjet
Southjet cargo
Southjet connect
Southwest Airlines
Spanair
SpiceJet
Spike Airlines
Spirit Airlines
Spirit of Manila Airlines
Spring Airlines
Spring Airlines Japan
Sprintair
SriLankan Airlines
Sriwijaya Air
Star Flyer
Star Peru (2I)
Star1 Airlines
Starbow Airlines
Starline.kz
Starlux Airlines
Sterling Airlines
STP Airways
Strategic Airlines
Sudan Airways
Sun Country Airlines
Sun D'Or
SunExpress
Sunrise Airways
Sunwing
Super Air Jet
SUR Lineas Aereas
Surinam Airways
SVG Air
Svyaz Rossiya
Swe Fly
Swiss European Air Lines
Swiss International Air Lines
Swissair
Syrian Arab Airlines
Syrian Pearl Airlines
T-way Air
T.J. Air
TAAG Angola Airlines
TACA Peru
TACV
Tajik Air
Tajikistan International Airlines
TAM Brazilian Airlines
TAM Mercosur
TAME
TAMPA
TAP Portugal
Tarom
Texas Wings
Tez Jet Airlines
Thai Air Cargo
Thai AirAsia
Thai Airways International
Thai Lion Air
Thai Smile Airways
Thai Vietjet Air
Thomas Cook Airlines
Thomsonfly
Tianjin Airlines
Tiara Air
Tiger Airways
Tiger Airways Australia
Tom\\'s & co airliners
Tomp Airlines
Tomsk-Avia
Trans Mediterranean Airlines
Trans Pas Air
Trans States Airlines
Transaero Airlines
Transair
TransAsia Airways
Transavia Denmark
Transavia France
Transavia Holland
TransBrasil Airlines
TransHolding
TransHolding System
Transilvania
Translift Airways
TransNusa Air
Transportes Aereos Cielos Andinos
TransRussiaAirlines
Transwest Air
TrasBrasil
Travel Service
Trigana Air Service
TRIP Linhas A
Tropic Air
TUIfly
TUIfly (X3)
TUIfly Nordic
Tuninter
Tunisair
Turan Air
Turk Hava Kurumu Hava Taksi Isletmesi
Turkish Air Force
Turkish Airlines
Turkish Wings Domestic
Turkmenistan Airlines
Turkuaz Airlines
Tway Airlines
Twin Jet
Tyrolean Airways
U.S. Air
Ukraine Atlantic
Ukraine International Airlines
UM Airlines
Uni Air
United Air Charters
United Airlines
United States Air Force
University of Birmingham Air Squadron (RAF)
Ural Airlines
US Airways
US Helicopter
US-Bangla Airlines
Usa Sky Cargo
USA3000 Airlines
UTair Aviation
UTair-Express
UVT Aero
Uzbekistan Airways
V Air
Valuair
Varig Log
Vasco Air
VASP
Via Conectia Airlines
VIA Líneas Aéreas
ViaAir
VickJet
VietJet Air
Vietnam Airlines
Viking Hellas
VIM Airlines
VIP Ecuador
Virgin America
Virgin Atlantic Airways
Virgin Australia
Virgin Express
Virgin Nigeria Airways
Virgin Pacific
Virginwings
Vision Air International
Vision Airlines (V2)
Viva Macau
VivaColombia
Vladivostok Air
VLM Airlines
Voestar
Volare Airlines
Volaris
Volga-Dnepr Airlines
Volotea
Volotea Costa Rica
VRG Linhas Aereas
Vuela Cuba
Vueling Airlines
Vuola Italia
Wataniya Airways
Wayraper
WebJet Linhas A
Welcome Air
West Air China
West Coast Air
Westair Aviation
Western Airlines
Westfalia Express VA
WestJet
WestJet Encore
Whitaker Air
Whitejets
Widerøe
Wilderness Air
Wind Jet
Wind Rose Aviation
Windward Islands Airways
Wings Air
Wizz Air
Wizz Air Hungary
Wizz Air Ukraine
World Airways
World Experience Airline
World Scale Airlines
XAIR USA
Xiamen Airlines
XiamenAir
XL Airways France
XOJET
Xpressair
XPTO
Yamal Airlines
Yangon Airways
Yellowstone Club Private Shuttle
Yellowtail
Yemenia
YES Airways
Yeti Airways
Yunnan Airlines
Yuzhmashavia
ZABAIKAL AIRLINES
Zabaykalskii Airlines
Zambezi Airlines (ZMA)
Zambia Skyways
Zanair
Zapolyarie Airlines
Zenith International Airline
Zest Air
Zoom Airlines
Zz
Аэросервис
Катэкавиа
Псковавиа"""
    
    airlines_list = [line.strip() for line in airlines.split('\n') if line.strip()]
    
    print(f"Всего авиакомпаний для проверки: {len(airlines_list)}\n")
    print("="*70)
    
    checker = AirlineChecker()
    results = []
    
    for idx, airline in enumerate(airlines_list, 1):
        print(f"\n[{idx}/{len(airlines_list)}] ", end='')
        result = checker.check_airline(airline)
        results.append(result)
        
        # Задержка между запросами (уменьшена)
        time.sleep(0.1)
        
        # Промежуточное сохранение каждые 100 компаний
        if idx % 100 == 0:
            temp_filename = f'airline_status_report_temp_{idx}.xlsx'
            checker.create_excel_report(results, temp_filename)
            print(f"\n\nПромежуточное сохранение: {idx} компаний проверено")
    
    # Финальное сохранение
    final_filename = 'airline_status_report_final.xlsx'
    checker.create_excel_report(results, final_filename)
    
    # Статистика
    print("\n" + "="*70)
    print("СТАТИСТИКА:")
    print(f"Всего проверено: {len(results)}")
    print(f"Информация найдена: {sum(1 for r in results if r['status'] != 'ИНФОРМАЦИЯ НЕ НАЙДЕНА')}")
    print(f"Информация не найдена: {sum(1 for r in results if r['status'] == 'ИНФОРМАЦИЯ НЕ НАЙДЕНА')}")
    print(f"Высокая уверенность: {sum(1 for r in results if r['confidence'] == 'ВЫСОКИЙ')}")
    print(f"Средняя уверенность: {sum(1 for r in results if r['confidence'] == 'СРЕДНИЙ')}")
    print(f"Низкая уверенность: {sum(1 for r in results if r['confidence'] == 'НИЗКИЙ')}")
    print(f"\nИтоговый отчет: {final_filename}")
    print("="*70)

if __name__ == "__main__":
    main()
